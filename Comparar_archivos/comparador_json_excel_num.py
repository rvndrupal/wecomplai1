"""
Comparador de Archivos JSON vs Excel
Autor: RVN - weComplai Qa Automiatizado
Descripción: Compara archivos JSON contra Excel tomando el Excel como base maestra
Versión 2: Maneja JSON con estructura anidada (referencias dentro de referencias)
"""

import pandas as pd
import json
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================================
# CONFIGURACIÓN
# ============================================================================

# 🎯 PARÁMETROS PRINCIPALES
NUMERO_FILAS_VALIDAR = 100  # Cambiar hasta 5000 datos
# NUMERO_FILAS_VALIDAR = None  # Descomentar para procesar completo

# 🔢 RANGO DE FILAS A PROCESAR (NUEVO)
FILA_INICIO = 1      # Fila donde iniciar (1 = primera fila de datos)
FILA_FIN = 100   # Fila donde terminar (hasta =5000)

# 📁 RUTAS DE ARCHIVOS
EXCEL_BASE_PATH = "mapeo_completo_02_di_038.xlsx" #EXCEL
JSON_PATH = "BASE_JSON.json"  # JSON
CARPETA_REPORTES = 'REPORTES_NUM'

# 🎨 COLORES PARA REPORTES
COLOR_VERDE = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
COLOR_ROJO = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
COLOR_AMARILLO = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
COLOR_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
FONT_HEADER = Font(bold=True, color="FFFFFF", size=11)
FONT_BOLD = Font(bold=True)

# ============================================================================
# FUNCIONES PRINCIPALES
# ============================================================================

def extraer_registros_recursivo(obj, registros=None):
    """
    Extrae recursivamente todos los registros que tengan 'secuencia'
    desde una estructura JSON anidada (maneja referencias dentro de referencias)
    
    Args:
        obj: Objeto JSON (dict o list)
        registros: Lista acumuladora de registros
    
    Returns:
        Lista con todos los registros encontrados
    """
    if registros is None:
        registros = []
    
    if isinstance(obj, dict):
        # Si el objeto tiene 'secuencia', es un registro válido
        if 'secuencia' in obj:
            # Crear una copia sin el campo 'referencias' para evitar duplicados
            registro_limpio = {k: v for k, v in obj.items() if k != 'referencias'}
            registros.append(registro_limpio)
            
            # Si tiene referencias, procesarlas recursivamente
            if 'referencias' in obj and isinstance(obj['referencias'], list):
                for ref in obj['referencias']:
                    extraer_registros_recursivo(ref, registros)
        else:
            # Si no tiene secuencia, buscar en sus valores
            for value in obj.values():
                extraer_registros_recursivo(value, registros)
    
    elif isinstance(obj, list):
        # Si es lista, procesar cada elemento
        for item in obj:
            extraer_registros_recursivo(item, registros)
    
    return registros


def cargar_excel(ruta, num_filas=None, fila_inicio=1, fila_fin=None):
    """
    Carga el archivo Excel base con opción de rango específico
    
    Args:
        ruta: Ruta del archivo Excel
        num_filas: Número de filas a procesar (None = todas)
        fila_inicio: Fila donde iniciar (1 = primera fila)
        fila_fin: Fila donde terminar (None = hasta el final)
    
    Returns:
        DataFrame con los datos del rango especificado
    """
    print(f"📂 Cargando Excel: {ruta}")
    
    # Cargar Excel completo primero
    df = pd.read_excel(ruta)
    total_filas_excel = len(df)
    
    # Aplicar rango de filas
    inicio_idx = fila_inicio - 1  # Convertir a índice 0-based
    fin_idx = fila_fin if fila_fin else total_filas_excel
    
    # Validar rangos
    if inicio_idx < 0:
        inicio_idx = 0
    if fin_idx > total_filas_excel:
        fin_idx = total_filas_excel
    
    # Extraer el rango
    df = df.iloc[inicio_idx:fin_idx]
    
    # Aplicar límite adicional si se especificó num_filas
    if num_filas and len(df) > num_filas:
        df = df.head(num_filas)
    
    filas_cargadas = len(df)
    
    print(f"   ✅ Cargadas {filas_cargadas} filas")
    print(f"   📊 Rango: Fila {fila_inicio} a {inicio_idx + filas_cargadas}")
    print(f"   📊 Total de columnas: {len(df.columns)}")
    
    return df


def cargar_json(ruta):
    """
    Carga el archivo JSON con los datos a comparar
    Extrae recursivamente todos los registros con 'secuencia'
    
    Args:
        ruta: Ruta del archivo JSON
    
    Returns:
        Lista con los registros JSON aplanados
    """
    print(f"📂 Cargando JSON: {ruta}")
    
    if not os.path.exists(ruta):
        print(f"   ⚠️  ADVERTENCIA: No se encontró el archivo JSON")
        print(f"   📝 Se esperaba en: {ruta}")
        return []
    
    with open(ruta, 'r', encoding='utf-8') as f:
        datos = json.load(f)
    
    print(f"   🔍 Extrayendo registros con 'secuencia' recursivamente...")
    
    # Extraer todos los registros con secuencia
    registros = extraer_registros_recursivo(datos)
    
    print(f"   ✅ Cargados {len(registros)} registros JSON")
    
    # Mostrar muestra de los primeros registros encontrados
    if len(registros) > 0:
        print(f"   📋 Primeras secuencias encontradas: ", end="")
        secuencias = [str(r.get('secuencia', '?')) for r in registros[:10]]
        print(", ".join(secuencias))
    
    return registros


def extraer_texto_puro(texto):
    """
    Extrae solo el texto puro de un JSON, eliminando caracteres estructurales
    
    Args:
        texto: String que puede contener JSON
    
    Returns:
        String con solo el texto interno
    """
    if not texto:
        return ""
    
    texto_str = str(texto)
    
    # Intentar parsear como JSON para extraer valores
    try:
        obj = json.loads(texto_str)
        
        # Función recursiva para extraer todos los valores de texto
        def extraer_valores(obj):
            valores = []
            if isinstance(obj, dict):
                for v in obj.values():
                    valores.extend(extraer_valores(v))
            elif isinstance(obj, list):
                for item in obj:
                    valores.extend(extraer_valores(item))
            else:
                valores.append(str(obj))
            return valores
        
        # Obtener todos los valores y concatenarlos
        valores = extraer_valores(obj)
        return ' '.join(valores)
    except:
        # Si no es JSON, devolver el texto limpiando caracteres estructurales
        import re
        # Eliminar caracteres JSON estructurales
        texto_limpio = re.sub(r'[{}\[\]":,]', ' ', texto_str)
        # Normalizar espacios
        texto_limpio = re.sub(r'\s+', ' ', texto_limpio).strip()
        return texto_limpio


def calcular_diferencia_texto(texto_excel, texto_json):
    """
    Calcula y muestra exactamente qué difiere entre dos textos
    Compara solo el contenido textual, ignorando caracteres estructurales JSON
    
    Args:
        texto_excel: Texto del Excel
        texto_json: Texto del JSON
    
    Returns:
        String describiendo la diferencia exacta
    """
    # Extraer solo el texto puro de ambos
    texto_puro_excel = extraer_texto_puro(texto_excel)
    texto_puro_json = extraer_texto_puro(texto_json)
    
    if texto_puro_excel == texto_puro_json:
        return ""
    
    # Si uno está vacío
    if not texto_puro_excel and texto_puro_json:
        return f"Falta en Excel: '{texto_puro_json}'"
    if texto_puro_excel and not texto_puro_json:
        return f"Falta en JSON: '{texto_puro_excel}'"
    
    # Si tienen diferente longitud
    if len(texto_puro_excel) != len(texto_puro_json):
        diferencia_long = abs(len(texto_puro_excel) - len(texto_puro_json))
        
        if len(texto_puro_excel) > len(texto_puro_json):
            # Encontrar qué tiene de más Excel
            texto_extra = texto_puro_excel.replace(texto_puro_json, '', 1)
            return f"Excel tiene más: '{texto_extra[:100]}...'" if len(texto_extra) > 100 else f"Excel tiene más: '{texto_extra}'"
        else:
            # Encontrar qué tiene de más JSON
            texto_extra = texto_puro_json.replace(texto_puro_excel, '', 1)
            return f"JSON tiene más: '{texto_extra[:100]}...'" if len(texto_extra) > 100 else f"JSON tiene más: '{texto_extra}'"
    
    # Si tienen la misma longitud pero contenido diferente
    # Encontrar las palabras que difieren
    palabras_excel = texto_puro_excel.split()
    palabras_json = texto_puro_json.split()
    
    diferencias = []
    for i, (p_excel, p_json) in enumerate(zip(palabras_excel, palabras_json)):
        if p_excel != p_json:
            diferencias.append(f"Palabra {i+1}: Excel='{p_excel}' vs JSON='{p_json}'")
            if len(diferencias) >= 2:
                break
    
    if diferencias:
        return " | ".join(diferencias)
    
    return f"Excel: '{texto_puro_excel[:50]}' vs JSON: '{texto_puro_json[:50]}'"


def normalizar_valor(valor, columna=None):
    """
    Normaliza valores para comparación
    - Convierte NaN a string vacío
    - Convierte números float a int cuando son enteros (1.0 → 1)
    - Convierte diccionarios y listas a JSON string normalizado
    - Para campos JSON específicos, parsea y compara como objetos
    - Normaliza espacios en blanco
    
    Args:
        valor: Valor a normalizar
        columna: Nombre de la columna (para tratamiento especial de campos JSON)
    
    Returns:
        String normalizado o objeto parseado
    """
    # Manejar listas y diccionarios PRIMERO (antes de pd.isna)
    if isinstance(valor, (list, dict)):
        # Convertir a JSON string sin espacios extras
        return json.dumps(valor, ensure_ascii=False, separators=(',', ':'), sort_keys=True).strip()
    
    # Ahora sí podemos verificar NaN de forma segura
    try:
        if pd.isna(valor):
            return ""
    except (ValueError, TypeError):
        # Si falla pd.isna, continuar con el valor
        pass
    
    # Normalizar números: convertir float a int si son enteros
    if isinstance(valor, (int, float)):
        # Si es un número decimal pero es entero (ej: 1.0, 2.0)
        if isinstance(valor, float) and valor.is_integer():
            valor = int(valor)
        return str(valor)
    
    # Campos que son JSON y deben compararse como objetos
    campos_json = [
        'requerimiento_especifico',
        'requerimiento_pregunta',
        'requerimiento_entregable',
        'ley_articulo_relacionado',
        'periodicidad_obligacion',
        'palabra_clave',  # Agregado
        'formula_tabla_anexo'  # Por si hay más campos JSON
    ]
    
    # Si es un campo JSON específico, intentar parsearlo
    if columna and columna in campos_json:
        if isinstance(valor, str) and valor.strip():
            try:
                # Intentar parsear el string JSON
                obj_parseado = json.loads(valor)
                # Volver a convertir a string normalizado
                return json.dumps(obj_parseado, ensure_ascii=False, separators=(',', ':'), sort_keys=True).strip()
            except:
                # Si no se puede parsear, devolver como string normalizado
                pass
    
    # Convertir a string y normalizar espacios
    valor_str = str(valor).strip()
    
    # Reemplazar múltiples espacios por uno solo
    import re
    valor_str = re.sub(r'\s+', ' ', valor_str)
    
    return valor_str


def buscar_registro_json(datos_json, secuencia, orden):
    """
    Busca un registro en el JSON por secuencia y orden
    
    Args:
        datos_json: Lista de registros JSON
        secuencia: Número de secuencia
        orden: Número de orden
    
    Returns:
        Diccionario con el registro encontrado o None
    """
    for registro in datos_json:
        if registro.get('secuencia') == secuencia and registro.get('orden') == orden:
            return registro
    return None


def comparar_campos(fila_excel, registro_json, columnas):
    """
    Compara campo por campo entre Excel y JSON
    
    Args:
        fila_excel: Serie de pandas con datos del Excel
        registro_json: Diccionario con datos del JSON
        columnas: Lista de columnas a comparar
    
    Returns:
        Lista de diccionarios con resultados de comparación
    """
    resultados = []
    
    for columna in columnas:
        valor_excel = normalizar_valor(fila_excel.get(columna), columna)
        
        # Obtener valor del JSON
        if columna in registro_json:
            valor_json = normalizar_valor(registro_json[columna], columna)
        else:
            valor_json = ""
        
        # Comparar
        coincide = (valor_excel == valor_json)
        
        # Calcular diferencia exacta si no coinciden
        diferencia_detallada = ""
        if not coincide:
            diferencia_detallada = calcular_diferencia_texto(valor_excel, valor_json)
        
        resultado = {
            'secuencia': fila_excel.get('secuencia'),
            'orden': fila_excel.get('orden'),
            'campo': columna,
            'valor_excel': valor_excel[:200] if len(str(valor_excel)) > 200 else valor_excel,  # Limitar longitud
            'valor_json': valor_json[:200] if len(str(valor_json)) > 200 else valor_json,
            'coincide': coincide,
            'diferencia': 'DIFERENTE' if not coincide else '',
            'detalle_diferencia': diferencia_detallada
        }
        
        resultados.append(resultado)
    
    return resultados


def procesar_comparacion(df_excel, datos_json):
    """
    Procesa la comparación completa entre Excel y JSON
    
    Args:
        df_excel: DataFrame con datos del Excel
        datos_json: Lista con registros JSON
    
    Returns:
        Tuple (resultados_detallados, resumen_por_fila)
    """
    print("\n🔄 Procesando comparación...")
    
    columnas = df_excel.columns.tolist()
    todos_resultados = []
    resumen_filas = []
    
    total_filas = len(df_excel)
    
    for idx, fila in df_excel.iterrows():
        secuencia = fila['secuencia']
        orden = fila['orden']
        
        print(f"   Procesando fila {idx + 1}/{total_filas} - Secuencia: {secuencia}, Orden: {orden}")
        
        # Buscar registro correspondiente en JSON
        registro_json = buscar_registro_json(datos_json, secuencia, orden)
        
        if registro_json is None:
            print(f"      ⚠️  No se encontró en JSON")
            resumen_filas.append({
                'secuencia': secuencia,
                'orden': orden,
                'encontrado_json': 'NO',
                'total_campos': len(columnas),
                'campos_coinciden': 0,
                'campos_difieren': len(columnas),
                'porcentaje_coincidencia': 0.0
            })
            continue
        
        # Comparar campos
        resultados_fila = comparar_campos(fila, registro_json, columnas)
        todos_resultados.extend(resultados_fila)
        
        # Calcular resumen de esta fila
        campos_coinciden = sum(1 for r in resultados_fila if r['coincide'])
        campos_difieren = len(columnas) - campos_coinciden
        porcentaje = (campos_coinciden / len(columnas)) * 100
        
        resumen_filas.append({
            'secuencia': secuencia,
            'orden': orden,
            'encontrado_json': 'SÍ',
            'total_campos': len(columnas),
            'campos_coinciden': campos_coinciden,
            'campos_difieren': campos_difieren,
            'porcentaje_coincidencia': round(porcentaje, 2)
        })
    
    print(f"   ✅ Comparación completada: {len(todos_resultados)} comparaciones realizadas")
    
    return todos_resultados, resumen_filas


def aplicar_formato_excel(ruta_excel, nombre_hoja, es_detallado=False):
    """
    Aplica formato condicional y estilos al Excel de forma ULTRA OPTIMIZADA
    Todos los colores se aplican pero de manera eficiente
    
    Args:
        ruta_excel: Ruta del archivo Excel
        nombre_hoja: Nombre de la hoja a formatear
        es_detallado: Si es el reporte detallado (aplica colores por coincidencia)
    """
    print(f"   🎨 Aplicando formato a hoja: {nombre_hoja}")
    
    wb = load_workbook(ruta_excel)
    ws = wb[nombre_hoja]
    
    # Pre-crear objetos de estilo (OPTIMIZACIÓN CLAVE)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    alignment_center = Alignment(horizontal='center', vertical='center')
    alignment_vertical = Alignment(vertical='center')
    
    # 1. FORMATEAR ENCABEZADOS
    for cell in ws[1]:
        cell.fill = COLOR_HEADER
        cell.font = FONT_HEADER
        cell.alignment = alignment_center
        cell.border = thin_border
    
    # 2. AJUSTAR ANCHO DE COLUMNAS (solo primeras 200 filas)
    print(f"      📏 Calculando ancho de columnas...")
    max_rows_for_width = min(200, ws.max_row)
    
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        
        for row_idx in range(1, max_rows_for_width + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            try:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 3. APLICAR COLORES DE FORMA OPTIMIZADA
    if es_detallado:
        print(f"      🎨 Aplicando colores (procesando por lotes)...")
        
        # Encontrar columna "coincide"
        coincide_col = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == 'coincide':
                coincide_col = idx
                break
        
        if coincide_col:
            total_rows = ws.max_row
            total_cols = ws.max_column
            
            # OPTIMIZACIÓN: Procesar por lotes de 500 filas
            batch_size = 500
            num_batches = (total_rows - 1) // batch_size + 1
            
            for batch_num in range(num_batches):
                start_row = 2 + (batch_num * batch_size)
                end_row = min(start_row + batch_size, total_rows + 1)
                
                print(f"         Lote {batch_num + 1}/{num_batches}: filas {start_row} a {end_row - 1}")
                
                for row_idx in range(start_row, end_row):
                    cell_coincide = ws.cell(row=row_idx, column=coincide_col)
                    
                    # Determinar color
                    if cell_coincide.value == True or cell_coincide.value == 'SÍ':
                        color = COLOR_VERDE
                    else:
                        color = COLOR_ROJO
                    
                    # Aplicar formato a toda la fila de forma eficiente
                    for col_idx in range(1, total_cols + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.fill = color
                        cell.border = thin_border
                        cell.alignment = alignment_vertical
    else:
        # Formato para reporte resumen
        print(f"      🎨 Aplicando bordes...")
        total_rows = ws.max_row
        total_cols = ws.max_column
        
        # Aplicar bordes por lotes
        batch_size = 500
        num_batches = (total_rows - 1) // batch_size + 1
        
        for batch_num in range(num_batches):
            start_row = 2 + (batch_num * batch_size)
            end_row = min(start_row + batch_size, total_rows + 1)
            
            for row_idx in range(start_row, end_row):
                for col_idx in range(1, total_cols + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = thin_border
                    cell.alignment = alignment_vertical
    
    # 4. CONGELAR PRIMERA FILA
    ws.freeze_panes = 'A2'
    
    print(f"      💾 Guardando cambios...")
    wb.save(ruta_excel)
    print(f"      ✅ Formato aplicado exitosamente")


def generar_reportes(resultados_detallados, resumen_filas):
    """
    Genera los reportes Excel con formato profesional
    
    Args:
        resultados_detallados: Lista con comparaciones detalladas
        resumen_filas: Lista con resumen por fila
    
    Returns:
        Tuple (ruta_resumen, ruta_detallado)
    """
    print("\n📊 Generando reportes...")
    
    # Formato de fecha: 08-12-2024_01-35PM
    timestamp = datetime.now().strftime("%d-%m-%Y_%I-%M%p")
    fecha_ejecucion = datetime.now().strftime("%d/%m/%Y - %I:%M %p")
    
    # Crear DataFrames
    df_detallado = pd.DataFrame(resultados_detallados)
    df_resumen = pd.DataFrame(resumen_filas)
    
    # Calcular estadísticas globales
    total_filas = len(resumen_filas)
    filas_encontradas = sum(1 for f in resumen_filas if f['encontrado_json'] == 'SÍ')
    filas_no_encontradas = total_filas - filas_encontradas
    
    if filas_encontradas > 0:
        total_campos = sum(f['campos_coinciden'] + f['campos_difieren'] for f in resumen_filas if f['encontrado_json'] == 'SÍ')
        campos_coinciden = sum(f['campos_coinciden'] for f in resumen_filas if f['encontrado_json'] == 'SÍ')
        campos_difieren = total_campos - campos_coinciden
        porcentaje_coincidencia = (campos_coinciden / total_campos) * 100 if total_campos > 0 else 0
    else:
        total_campos = campos_coinciden = campos_difieren = 0
        porcentaje_coincidencia = 0
    
    # Contar filas con 100% coincidencia vs filas con diferencias
    filas_100_coincidencia = sum(1 for f in resumen_filas if f.get('porcentaje_coincidencia', 0) == 100.0)
    filas_con_diferencias = filas_encontradas - filas_100_coincidencia
    
    # Rutas de archivos con nuevo formato
    ruta_resumen = os.path.join(CARPETA_REPORTES, f'RESUMEN_Comparacion_{timestamp}.xlsx')
    ruta_detallado = os.path.join(CARPETA_REPORTES, f'DETALLADO_Comparacion_{timestamp}.xlsx')
    
    # Crear carpeta si no existe
    os.makedirs(CARPETA_REPORTES, exist_ok=True)
    
    # Crear hoja de información
    modo_ejecucion = ''
    if NUMERO_FILAS_VALIDAR:
        modo_ejecucion = f'VALIDACIÓN ({NUMERO_FILAS_VALIDAR} filas)'
    elif FILA_FIN:
        modo_ejecucion = f'RANGO (Filas {FILA_INICIO} a {FILA_FIN})'
    else:
        modo_ejecucion = f'COMPLETO (Desde fila {FILA_INICIO})'
    
    info_data = {
        'Información': [
            'Fecha de Ejecución',
            'Archivo Excel Base',
            'Archivo JSON',
            'Fila Inicio',
            'Fila Fin',
            'Total Filas Procesadas',
            'Modo de Ejecución',
            'Autor'
        ],
        'Valor': [
            fecha_ejecucion,
            EXCEL_BASE_PATH,
            JSON_PATH,
            FILA_INICIO,
            FILA_FIN if FILA_FIN else 'Hasta el final',
            len(resumen_filas),
            modo_ejecucion,
            'RVN - weComplai Qa Automiatizado'
        ]
    }
    df_info = pd.DataFrame(info_data)
    
    # Obtener lista de filas con diferencias (secuencia-orden)
    filas_con_dif_list = [f"{f['secuencia']}-{f['orden']}" for f in resumen_filas if f.get('porcentaje_coincidencia', 0) < 100.0 and f['encontrado_json'] == 'SÍ']
    filas_con_dif_texto = ', '.join(filas_con_dif_list) if filas_con_dif_list else 'Ninguna'
    
    # Crear hoja de estadísticas
    stats_data = {
        'Métrica': [
            '📊 FILAS PROCESADAS',
            'Total de filas procesadas',
            'Filas encontradas en JSON',
            'Filas NO encontradas en JSON',
            'Filas con 100% coincidencia',
            'Filas con diferencias',
            'Detalle de filas con diferencias',
            '',
            '📈 CAMPOS COMPARADOS',
            'Total de comparaciones',
            'Campos que coinciden',
            'Campos diferentes',
            'Porcentaje de coincidencia',
            '',
            '✅ RESULTADO',
            'Estado general'
        ],
        'Valor': [
            '',
            total_filas,
            filas_encontradas,
            filas_no_encontradas,
            filas_100_coincidencia,
            filas_con_diferencias,
            filas_con_dif_texto,
            '',
            '',
            total_campos,
            campos_coinciden,
            campos_difieren,
            f'{porcentaje_coincidencia:.2f}%',
            '',
            '',
            'APROBADO ✅' if filas_con_diferencias == 0 else f'RECHAZADO ❌ ({filas_con_diferencias} filas con diferencias)'
        ]
    }
    df_stats = pd.DataFrame(stats_data)
    
    # Guardar reportes con hojas de información y estadísticas
    print(f"   📝 Guardando reporte resumen...")
    with pd.ExcelWriter(ruta_resumen, engine='openpyxl') as writer:
        df_info.to_excel(writer, sheet_name='Info', index=False)
        df_stats.to_excel(writer, sheet_name='Estadísticas', index=False)
        df_resumen.to_excel(writer, sheet_name='Resumen', index=False)
    aplicar_formato_excel(ruta_resumen, 'Resumen', es_detallado=False)
    aplicar_formato_excel(ruta_resumen, 'Info', es_detallado=False)
    aplicar_formato_excel(ruta_resumen, 'Estadísticas', es_detallado=False)
    
    print(f"   📝 Guardando reporte detallado...")
    with pd.ExcelWriter(ruta_detallado, engine='openpyxl') as writer:
        df_info.to_excel(writer, sheet_name='Info', index=False)
        df_stats.to_excel(writer, sheet_name='Estadísticas', index=False)
        df_detallado.to_excel(writer, sheet_name='Detallado', index=False)
    aplicar_formato_excel(ruta_detallado, 'Detallado', es_detallado=True)
    aplicar_formato_excel(ruta_detallado, 'Info', es_detallado=False)
    aplicar_formato_excel(ruta_detallado, 'Estadísticas', es_detallado=False)
    
    print(f"\n✅ Reportes generados exitosamente:")
    print(f"   📄 Resumen: {os.path.basename(ruta_resumen)}")
    print(f"   📄 Detallado: {os.path.basename(ruta_detallado)}")
    print(f"   📅 Fecha: {fecha_ejecucion}")
    print(f"   📊 Filas con 100% coincidencia: {filas_100_coincidencia}/{total_filas}")
    print(f"   ⚠️  Filas con diferencias: {filas_con_diferencias}/{total_filas}")
    
    return ruta_resumen, ruta_detallado


def mostrar_estadisticas(resumen_filas):
    """
    Muestra estadísticas generales de la comparación
    
    Args:
        resumen_filas: Lista con resumen por fila
    """
    print("\n" + "=" * 80)
    print("📊 ESTADÍSTICAS GENERALES")
    print("=" * 80)
    
    total_filas = len(resumen_filas)
    filas_encontradas = sum(1 for f in resumen_filas if f['encontrado_json'] == 'SÍ')
    filas_no_encontradas = total_filas - filas_encontradas
    
    if filas_encontradas > 0:
        total_campos = sum(f['campos_coinciden'] + f['campos_difieren'] for f in resumen_filas if f['encontrado_json'] == 'SÍ')
        campos_coinciden = sum(f['campos_coinciden'] for f in resumen_filas if f['encontrado_json'] == 'SÍ')
        campos_difieren = total_campos - campos_coinciden
        porcentaje_general = (campos_coinciden / total_campos) * 100 if total_campos > 0 else 0
    else:
        total_campos = campos_coinciden = campos_difieren = 0
        porcentaje_general = 0
    
    print(f"\n📋 Resumen de Filas:")
    print(f"   • Total de filas procesadas: {total_filas}")
    print(f"   • Filas encontradas en JSON: {filas_encontradas}")
    print(f"   • Filas NO encontradas en JSON: {filas_no_encontradas}")
    
    print(f"\n📊 Resumen de Campos:")
    print(f"   • Total de comparaciones: {total_campos}")
    print(f"   • Campos que coinciden: {campos_coinciden} ({porcentaje_general:.2f}%)")
    print(f"   • Campos diferentes: {campos_difieren} ({100-porcentaje_general:.2f}%)")
    
    print("\n" + "=" * 80)


# ============================================================================
# FUNCIÓN PRINCIPAL
# ============================================================================

def main():
    """
    Función principal que ejecuta todo el proceso
    """
    print("\n" + "=" * 80)
    print("🚀 COMPARADOR JSON vs EXCEL - RVN weComplai Qa Automiatizado")
    print("=" * 80)
    print(f"📅 Fecha de ejecución: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    # Mostrar modo de ejecución
    if NUMERO_FILAS_VALIDAR:
        print(f"⚙️  Modo: VALIDACIÓN - Procesando primeras {NUMERO_FILAS_VALIDAR} filas")
    else:
        if FILA_FIN:
            print(f"⚙️  Modo: RANGO - Procesando filas {FILA_INICIO} a {FILA_FIN}")
        else:
            print(f"⚙️  Modo: COMPLETO - Desde fila {FILA_INICIO} hasta el final")
    
    print("=" * 80)
    
    # 1. Cargar archivos
    df_excel = cargar_excel(EXCEL_BASE_PATH, NUMERO_FILAS_VALIDAR, FILA_INICIO, FILA_FIN)
    datos_json = cargar_json(JSON_PATH)
    
    if len(datos_json) == 0:
        print("\n❌ ERROR: No se pudo cargar el archivo JSON")
        print("   Por favor verifica que el archivo exista en la ruta correcta")
        return
    
    # 2. Procesar comparación
    resultados_detallados, resumen_filas = procesar_comparacion(df_excel, datos_json)
    
    # 3. Generar reportes
    ruta_resumen, ruta_detallado = generar_reportes(resultados_detallados, resumen_filas)
    
    # 4. Mostrar estadísticas
    mostrar_estadisticas(resumen_filas)
    
    print("\n✅ Proceso completado exitosamente!")
    print(f"📁 Los reportes están disponibles en: {CARPETA_REPORTES}")
    print("=" * 80 + "\n")


if __name__ == "__main__":
    main()

...